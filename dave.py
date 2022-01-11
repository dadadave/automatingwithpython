"""int thisa first pat of my code am editing the excel file and how is it done
 first i create a work book and i assigne the workbook to my excel file
 with the use of a for loop i will iterate throught the third column which is the email column
 while iterating i will use the replace funtion to change the helpinghands.cm to handsinhands.org
 an i will save the changes in a new file called updated_employee.xlsx
 """
import csv
from openpyxl import workbook, load_workbook #creating and loading a workbook
#creating a variable for the workbook and attributing it to the excel database
wb = load_workbook('employeedata.xlsx')
ws = wb.active
#iterating through the third column and changing the domain name to handsinhands.org
for i in range(2,ws.max_row + 1):
    cell = ws.cell(i,3)
    if 'helpinghands.cm' in cell.value:
        updated_email=(cell.value).replace('helpinghands.cm','handsinhands.org')
        ws.cell(i,3).value = updated_email
#saving the changes done in the file in a new excel file
wb.save('updated_employeedata.xlsx')
"""  in the second part of the code i will edit the csv file 
i will file import the csv libary the i will create the csv file for the updated emails in a writable format
the i open the employee.csv in a readable formate with the use of a for loop i will iterate throught the csv file and while iterating i 
wil still use the replace method and finally create a variable line and use the formate method to just say how the new csv file will be 
organised"""
import csv
#creating a new csv file where all changes will be saved
outfile = open('updated_employeedata.csv ', 'w')
#opening the database
with open('employeedata.csv', 'r') as csv_file:
    #creating a reading var
    csv_reader = csv.reader(csv_file)
    header = next(csv_reader)
    #iterating through the different rows
    for row in csv_reader:
        name = row[0]
        contact =row[1]
        #while iterating through the third column i change the domain name once
        email = row[2]
        address =row[3]
        if 'helpinghands.cm' in email:
            upemail=(email).replace('helpinghands.cm','handsinhands.org')
        line = "{},{},{},{}\n".format(name, contact,upemail, address)
        outfile.write(line)
outfile.close()


  

